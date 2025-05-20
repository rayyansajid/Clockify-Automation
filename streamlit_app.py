import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from io import BytesIO
import base64
from PIL import Image
from pathlib import Path

st.set_page_config(page_title="Clockify Timesheet Formatter - ei1", layout="wide")

# Load your logo image
logo_path = Path(__file__).parent / "ei1-logo.png"
logo = Image.open(logo_path)
# Convert to base64
buffered = BytesIO()
logo.save(buffered, format="PNG")
logo_base64 = base64.b64encode(buffered.getvalue()).decode()

# Create two columns
col1, col2 = st.columns([8, 2])  # Adjust width ratio as needed

with col2:
    st.markdown(
        f"""
        <div style="background-color: black; border-radius: 5px; text-align: center;">
            <img src="data:image/png;base64,{logo_base64}" width="300" />
        </div>
        """,
        unsafe_allow_html=True
    )
 # Adjust width to fit well with title

with col1:
    st.title("🕒 Clockify Timesheet Formatter")

# st.title("🕒 Clockify Timesheet Formatter")

uploaded_file = st.file_uploader("Upload your Clockify Excel file", type=["xlsx"])

if uploaded_file:
    with st.spinner("Processing file..."):
        df = pd.read_excel(uploaded_file)

        # Extract user info
        user_name = df['User'].iloc[0]
        user_email = df['Email'].iloc[0]

        # Columns to keep and reorder
        columns_to_keep = ['Project', 'Client', 'Description', 'Task', 'Start Date', 'Start Time', 'End Time', 'Duration (h)']
        desired_order = ['Day', 'Start Date', 'Start Time', 'End Time', 'Duration (h)', 'Project', 'Client', 'Description', 'Task']

        df = df[columns_to_keep]
        df['Day'] = pd.to_datetime(df['Start Date']).dt.day_name()
        df.sort_values(by=['Start Date', 'Start Time'], inplace=True)
        df = df[desired_order]

        # Save to a temporary buffer
        temp_buffer = BytesIO()
        df.to_excel(temp_buffer, index=False)
        temp_buffer.seek(0)

        wb = load_workbook(temp_buffer)
        ws = wb.active

        # Styles
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(bold=True, color="000000")
        light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        bold_font = Font(bold=True)
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        # Style header
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thick_border

        def parse_duration(duration_str):
            if isinstance(duration_str, str):
                parts = list(map(int, duration_str.split(':')))
                while len(parts) < 3:
                    parts.append(0)
                return timedelta(hours=parts[0], minutes=parts[1], seconds=parts[2])
            return timedelta()

        # Format by day
        row = 2
        weekly_duration = timedelta()
        while row <= ws.max_row:
            day = ws.cell(row=row, column=1).value
            start_row = row
            while row <= ws.max_row and ws.cell(row, column=1).value == day:
                row += 1
            end_row = row - 1

            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(start_row, 1).alignment = center_align
            start_date_val = ws.cell(start_row, 2).value
            if isinstance(start_date_val, datetime):
                ws.cell(start_row, 2).value = start_date_val.strftime("%d/%m/%Y")
            ws.cell(start_row, 2).alignment = center_align

            total_duration = timedelta()
            for r in range(start_row, end_row + 1):
                parsed = parse_duration(ws.cell(r, 5).value)
                total_duration += parsed
            weekly_duration += total_duration

            ws.insert_rows(row)
            ws.cell(row, 3).value = "Total"
            ws.cell(row, 3).fill = light_blue
            ws.cell(row, 3).font = bold_font
            ws.cell(row, 4).fill = light_blue
            ws.cell(row, 5).value = str(total_duration)
            ws.cell(row, 5).fill = light_blue
            ws.cell(row, 5).font = bold_font
            row += 1

        first_col = 1
        last_col = ws.max_column
        last_row = ws.max_row

        for r in range(1, last_row + 1):
            cell1 = ws.cell(r, 1)
            cell2 = ws.cell(r, 2)
            cell1.font = bold_font
            cell2.font = bold_font


        for r in range(1, last_row + 1):
            for c in range(first_col, last_col + 1):
                cell = ws.cell(r, c)
                cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=cell.border.top, bottom=cell.border.bottom)
                if c < 6:
                    cell.alignment = center_align

        date_to_last_row = {}
        for r in range(2, last_row + 1):
            val = ws.cell(r, 2).value
            if isinstance(val, str):
                try:
                    parsed = datetime.strptime(val, "%d/%m/%Y").date()
                    date_to_last_row[parsed] = r
                except:
                    continue

        for r in date_to_last_row.values():
            for c in range(first_col, last_col + 1):
                ws.cell(r - 1, c).border = Border(
                    left=ws.cell(r - 1, c).border.left,
                    right=ws.cell(r - 1, c).border.right,
                    top=ws.cell(r - 1, c).border.top,
                    bottom=Side(style='thick')
                )

        for r in range(2, last_row + 1):
            ws.cell(r, first_col).border = Border(
                left=Side(style='thick'), right=ws.cell(r, first_col).border.right,
                top=ws.cell(r, first_col).border.top, bottom=ws.cell(r, first_col).border.bottom
            )
            ws.cell(r, last_col).border = Border(
                right=Side(style='thick'), left=ws.cell(r, last_col).border.left,
                top=ws.cell(r, last_col).border.top, bottom=ws.cell(r, last_col).border.bottom
            )

        for c in range(first_col, last_col + 1):
            ws.cell(last_row, c).border = Border(
                left=ws.cell(last_row, c).border.left,
                right=ws.cell(last_row, c).border.right,
                top=ws.cell(last_row, c).border.top,
                bottom=Side(style='thick')
            )

        for r in range(2, last_row + 1):
            if ws.cell(r, 6).value == "Break":
                for c in range(6, last_col + 1):
                    ws.cell(r, c).font = bold_font

        def format_timedelta_hhmmss(td):
            total_seconds = int(td.total_seconds())
            h, rem = divmod(total_seconds, 3600)
            m, s = divmod(rem, 60)
            return f"{h:02}:{m:02}:{s:02}"

        ws.cell(last_row + 2, 3).value = "Total Hours in Week"
        ws.cell(last_row + 2, 3).fill = light_blue
        ws.cell(last_row + 2, 3).font = bold_font
        ws.cell(last_row + 2, 4).fill = light_blue
        ws.cell(last_row + 2, 5).value = format_timedelta_hhmmss(weekly_duration)
        ws.cell(last_row + 2, 5).font = bold_font
        ws.cell(last_row + 2, 5).fill = light_blue

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        # Save to in-memory buffer
        final_output = BytesIO()
        # Clean user name (remove spaces)
        user_clean = user_name.replace(" ", "")
        
        #########################################################################
        # Extract date range
        # start_date = pd.to_datetime(df['Start Date'].min()).strftime('%d%b')
        # end_date = pd.to_datetime(df['Start Date'].max()).strftime('%d%b')
        # year = pd.to_datetime(df['Start Date'].max()).year
        
        # # Construct final output filename
        # filename = f"Timesheet-{user_clean}-{start_date}-{end_date}-{year}.xlsx"
        # wb.save(final_output)
        
        # print(f"✅ Timesheet saved as: {filename}")
        #########################################################################       
        # Extract and format start and end dates
        start_dt = pd.to_datetime(df['Start Date'].min())
        end_dt = pd.to_datetime(df['Start Date'].max())
        
        start_str = start_dt.strftime("%d_%m_%Y")
        end_str = end_dt.strftime("%d_%m_%Y")
        
        # Construct the final filename
        filename = f"Timesheet-{user_clean}-{start_str}-{end_str}.xlsx"
        wb.save(final_output)
        
        print(f"✅ Timesheet saved as: {filename}")

        #########################################################################
        
        final_output.seek(0)

    st.success("✅ File processed successfully!")

    st.download_button(
        label="📥 Download Formatted Excel File",
        data=final_output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Add spacing
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center; font-size: 16px; padding-top: 20px;">
        Precision. Simplicity & Intellectuality Powered by <b>Team C</b>:<br> <i>Rayyan Sajid</i> & <i>Hamza Ahmed</i>
    </div>
    """,
    unsafe_allow_html=True
)
